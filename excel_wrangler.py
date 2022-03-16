# Calvin Crosby
# Excel Controller Class
from openpyxl import Workbook
import openpyxl
from interpreter import Interpreter

class ExcelWrangler():
    def __init__(self,spreadsheet) -> None:
        self.wb = spreadsheet
        self.sheet = self.wb.active
        self.interpreter = Interpreter() 
        self.general_info_hook = None
        self.selected_market_data_hook = None
        self.balance_sheet_hook = None
        self.cash_flow_statement_hook = None
        self.reported_income_statement_hook = None
        self.adjusted_income_statement_hook = None
        
        self.find_sheet_componenets()
    
    def find_sheet_componenets(self):
        '''This method identifies the locations of CCA Model's components. Once the've been
            identified, their values can be dynamically updated with the correct data, should
            small changes be made to the spreadsheet in the future.'''

        for row in self.sheet.iter_rows(min_row = 4, max_col = 20, max_row = 64):
            for cell in row:
                if cell.value == "Balance Sheet Data":
                    self.balance_sheet_hook = cell
                elif cell.value == "Reported Income Statement":
                    self.reported_income_statement_hook = cell
                elif cell.value == "General Information":
                    self.general_info_hook = cell
                elif cell.value == "Selected Market Data":
                    self.selected_market_data_hook = cell
                elif cell.value == "Cash Flow Statement Data":
                    self.cash_flow_statement_hook = cell
    def print_section_hooks(self):
        print("general info cell: " + str(self.general_info_hook))
        print("selected market data cell: " + str(self.selected_market_data_hook))
        print("balance sheet cell: " + str(self.balance_sheet_hook))
        print("cash flow statement cell: " + str(self.cash_flow_statement_hook))
        print("income statement cell: " + str(self.reported_income_statement_hook))
    
    def insert_balance_sheet(self,data):
        cur_year = data[0] # The most recent year of balance sheet data
        past_year = data[1] # The previous year's data
 
       
        cur_cell = self.balance_sheet_hook
        cur_cell = self.sheet.cell(row = cur_cell.row+1, column = cur_cell.column)
        print(cur_cell.value)
        print(cur_cell.row)
        print(cur_cell.column)
        cur_row = cur_cell.row
        cur_col = cur_cell.column

        done = False # will update to True once all  balance sheet items inserted

        while not done:
            cur_cell = self.sheet.cell(row = cur_row, column = cur_col)
            print(cur_cell.value)
    
            if cur_cell.value is None:
                # if the cell is empty, move down past it 
                cur_cell = self.sheet.cell(row = (cur_row+1), column = cur_col)
                cur_row+=1
                continue
            
            elif cur_cell.value in self.interpreter.balance_items:
                # this cell has a line item for the balance sheet. Send it off to the 
                # intepreter to get the old and current year values and populate the spreadsheet
                prior_yr_val = self.interpreter.interpret_value(cur_cell.value,past_year)
                cur_yr_val = self.interpreter.interpret_value(cur_cell.value,cur_year)

                self.sheet.cell(row = cur_row , column = (cur_col+3)).value = prior_yr_val
             
                self.sheet.cell(row = cur_row, column = (cur_col+4)).value = cur_yr_val

               # prior_cell.value = prior_yr_val
                #cur_cell.value = cur_yr_val

                if cur_cell.value == self.interpreter.balance_items[-1]: # the last line item, break out
                    done = True

                cur_row +=1
            else:
                cur_row +=1
                if cur_cell.row > 65:
                    break
            print(cur_cell)


            
            










def main():
    file = "cca_test.xlsx"

    wb = openpyxl.load_workbook(file,read_only=True)
    print(wb.sheetnames)
    sheet = wb['Sheet1']
    #for row in sheet.iter_rows(min_row = 4, max_col = 20, max_row = 64):
      #  for cell in row:
         #   print(cell.value)
   # print(sheet)
    wrangle = ExcelWrangler(wb)
    wrangle.find_sheet_componenets()
    wrangle.print_section_hooks()

if __name__ == "__main__":
    main()
